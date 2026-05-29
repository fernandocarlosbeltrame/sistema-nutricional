import sys
import json
from html import escape
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from PySide6.QtCore import QMarginsF, Qt
from PySide6.QtGui import QColor, QPageSize, QPdfWriter, QTextDocument
from PySide6.QtWidgets import (
    QApplication,
    QAbstractSpinBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSpinBox,
    QDoubleSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QVBoxLayout,
    QWidget,
)


APP_DIR = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
ALIMENTOS_PERSONALIZADOS_PATH = APP_DIR / "alimentos_personalizados.json"

ALIMENTOS_BASE = {
    "Arroz cozido": {"grupo": "Carboidrato", "faixa": "Econômica", "porcao": "100g", "calorias": 130, "proteina": 2.7, "carbo": 28, "gordura": 0.3, "preco": 0.80},
    "Feijão cozido": {"grupo": "Carboidrato", "faixa": "Econômica", "porcao": "100g", "calorias": 76, "proteina": 4.8, "carbo": 13.6, "gordura": 0.5, "preco": 0.90},
    "Macarrão cozido": {"grupo": "Carboidrato", "faixa": "Econômica", "porcao": "100g", "calorias": 157, "proteina": 5.8, "carbo": 30.9, "gordura": 0.9, "preco": 0.90},
    "Batata cozida": {"grupo": "Carboidrato", "faixa": "Média", "porcao": "100g", "calorias": 86, "proteina": 1.7, "carbo": 20, "gordura": 0.1, "preco": 0.85},
    "Pão francês": {"grupo": "Carboidrato", "faixa": "Econômica", "porcao": "1 unidade", "calorias": 135, "proteina": 4.5, "carbo": 28, "gordura": 1.5, "preco": 0.90},
    "Pão integral": {"grupo": "Carboidrato", "faixa": "Média", "porcao": "2 fatias", "calorias": 140, "proteina": 6, "carbo": 24, "gordura": 2.5, "preco": 1.60},
    "Aveia": {"grupo": "Carboidrato", "faixa": "Média", "porcao": "30g", "calorias": 117, "proteina": 5, "carbo": 20, "gordura": 2.1, "preco": 0.75},
    "Banana": {"grupo": "Fruta", "faixa": "Econômica", "porcao": "1 unidade", "calorias": 90, "proteina": 1.1, "carbo": 23, "gordura": 0.3, "preco": 0.80},
    "Maçã": {"grupo": "Fruta", "faixa": "Média", "porcao": "1 unidade", "calorias": 80, "proteina": 0.4, "carbo": 21, "gordura": 0.2, "preco": 1.30},
    "Mamão": {"grupo": "Fruta", "faixa": "Média", "porcao": "1 fatia", "calorias": 70, "proteina": 0.6, "carbo": 18, "gordura": 0.2, "preco": 1.20},
    "Frutas vermelhas": {"grupo": "Fruta", "faixa": "Premium", "porcao": "100g", "calorias": 60, "proteina": 1, "carbo": 14, "gordura": 0.3, "preco": 4.50},
    "Ovo": {"grupo": "Proteína", "faixa": "Econômica", "porcao": "1 unidade", "calorias": 70, "proteina": 6, "carbo": 0.5, "gordura": 5, "preco": 0.90},
    "Frango grelhado": {"grupo": "Proteína", "faixa": "Econômica", "porcao": "100g", "calorias": 165, "proteina": 31, "carbo": 0, "gordura": 3.6, "preco": 3.20},
    "Carne moída": {"grupo": "Proteína", "faixa": "Média", "porcao": "100g", "calorias": 250, "proteina": 26, "carbo": 0, "gordura": 15, "preco": 4.50},
    "Carne bovina": {"grupo": "Proteína", "faixa": "Média", "porcao": "100g", "calorias": 220, "proteina": 29, "carbo": 0, "gordura": 11, "preco": 6.50},
    "Linguiça": {"grupo": "Proteína", "faixa": "Média", "porcao": "100g", "calorias": 300, "proteina": 16, "carbo": 2, "gordura": 25, "preco": 4.00},
    "Carne suína": {"grupo": "Proteína", "faixa": "Média", "porcao": "100g", "calorias": 242, "proteina": 27, "carbo": 0, "gordura": 14, "preco": 5.20},
    "Peixe": {"grupo": "Proteína", "faixa": "Premium", "porcao": "100g", "calorias": 140, "proteina": 26, "carbo": 0, "gordura": 4, "preco": 5.50},
    "Leite": {"grupo": "Proteína", "faixa": "Econômica", "porcao": "200ml", "calorias": 120, "proteina": 6, "carbo": 9.5, "gordura": 6, "preco": 1.20},
    "Iogurte natural": {"grupo": "Proteína", "faixa": "Média", "porcao": "170g", "calorias": 110, "proteina": 7, "carbo": 12, "gordura": 3, "preco": 2.50},
    "Whey protein": {"grupo": "Proteína", "faixa": "Premium", "porcao": "30g", "calorias": 120, "proteina": 24, "carbo": 3, "gordura": 2, "preco": 4.00},
    "Azeite": {"grupo": "Gordura", "faixa": "Premium", "porcao": "1 colher", "calorias": 90, "proteina": 0, "carbo": 0, "gordura": 10, "preco": 0.70},
    "Pasta de amendoim": {"grupo": "Gordura", "faixa": "Média", "porcao": "1 colher", "calorias": 95, "proteina": 4, "carbo": 3, "gordura": 8, "preco": 0.90},
    "Castanhas": {"grupo": "Gordura", "faixa": "Premium", "porcao": "20g", "calorias": 120, "proteina": 4, "carbo": 4, "gordura": 10, "preco": 2.00},
    "Salada": {"grupo": "Legume/Verdura", "faixa": "Econômica", "porcao": "à vontade", "calorias": 30, "proteina": 1, "carbo": 6, "gordura": 0, "preco": 1.50},
    "Legumes": {"grupo": "Legume/Verdura", "faixa": "Média", "porcao": "100g", "calorias": 45, "proteina": 2, "carbo": 9, "gordura": 0.3, "preco": 1.50},
    "Brócolis": {"grupo": "Legume/Verdura", "faixa": "Premium", "porcao": "100g", "calorias": 35, "proteina": 2.4, "carbo": 7, "gordura": 0.4, "preco": 2.50},
}

REFEICOES = [
    "Café da manhã",
    "Almoço",
    "Lanche da tarde",
    "Janta",
    "Lanche da noite",
]

MISTURAS = {
    "ovo",
    "frango grelhado",
    "carne moída",
    "carne bovina",
    "linguiça",
    "carne suína",
    "peixe",
}


def carregar_alimentos():
    alimentos = {nome: dados.copy() for nome, dados in ALIMENTOS_BASE.items()}
    if ALIMENTOS_PERSONALIZADOS_PATH.exists():
        try:
            with open(ALIMENTOS_PERSONALIZADOS_PATH, "r", encoding="utf-8") as arquivo:
                personalizados = json.load(arquivo)
            for nome, dados in personalizados.items():
                if isinstance(dados, dict):
                    alimentos[nome] = dados
        except (OSError, json.JSONDecodeError):
            pass
    return alimentos


def salvar_alimentos_personalizados(alimentos):
    personalizados = {
        nome: dados
        for nome, dados in alimentos.items()
        if nome not in ALIMENTOS_BASE or dados != ALIMENTOS_BASE[nome]
    }
    with open(ALIMENTOS_PERSONALIZADOS_PATH, "w", encoding="utf-8") as arquivo:
        json.dump(personalizados, arquivo, ensure_ascii=False, indent=2)


def calcular_imc(peso, altura_cm):
    altura_m = altura_cm / 100
    return peso / (altura_m ** 2)


def classificar_imc(imc):
    if imc < 18.5:
        return "Abaixo do peso"
    if imc < 25:
        return "Peso normal"
    if imc < 30:
        return "Sobrepeso"
    return "Obesidade"


def calcular_calorias(peso, objetivo):
    if objetivo == "Emagrecer":
        return peso * 25
    if objetivo == "Ganhar massa":
        return peso * 35
    return peso * 30


def calcular_macros(peso, objetivo):
    if objetivo == "Emagrecer":
        return round(peso * 1.8), round(peso * 2.5), round(peso * 0.8)
    if objetivo == "Ganhar massa":
        return round(peso * 2.0), round(peso * 4.5), round(peso * 1.0)
    return round(peso * 1.6), round(peso * 3.5), round(peso * 0.9)


def multiplicador_refeicao(refeicao, objetivo):
    base = {
        "Café da manhã": 1.0,
        "Almoço": 1.4,
        "Lanche da tarde": 0.8,
        "Janta": 1.1,
        "Lanche da noite": 0.6,
    }
    if objetivo == "Emagrecer":
        return base[refeicao] * 0.8
    if objetivo == "Ganhar massa":
        return base[refeicao] * 1.3
    return base[refeicao]


def estruturas_refeicao(refeicao, mult):
    if refeicao == "Café da manhã":
        return [
            ("Proteína", ["Proteína"], 1.0 * mult),
            ("Carboidrato/Fruta", ["Carboidrato", "Fruta"], 1.0 * mult),
            ("Complemento", ["Gordura"], 0.5 * mult),
        ]
    if refeicao == "Almoço":
        return [
            ("Carboidrato", ["Carboidrato"], 1.4 * mult),
            ("Proteína", ["Proteína"], 1.2 * mult),
            ("Legumes/Verduras", ["Legume/Verdura"], 1.0 * mult),
        ]
    if refeicao == "Lanche da tarde":
        return [
            ("Fruta/Carboidrato", ["Fruta", "Carboidrato"], 1.0 * mult),
            ("Proteína", ["Proteína"], 0.8 * mult),
        ]
    if refeicao == "Janta":
        return [
            ("Proteína", ["Proteína"], 1.1 * mult),
            ("Carboidrato", ["Carboidrato"], 0.9 * mult),
            ("Legumes/Verduras", ["Legume/Verdura"], 1.0 * mult),
        ]
    return [
        ("Proteína leve", ["Proteína"], 0.6 * mult),
        ("Fruta/Gordura", ["Fruta", "Gordura"], 0.5 * mult),
    ]


def eh_mistura(nome, dados):
    return dados["grupo"] == "Proteína" and nome.lower() in MISTURAS


def escolher_alimento(alimentos, grupos, faixa, refeicao=None, tipo=None, usados=None):
    filtrados = [
        (nome, dados)
        for nome, dados in alimentos.items()
        if dados["grupo"] in grupos and (faixa == "Todas" or dados["faixa"] == faixa)
    ]
    if not filtrados and faixa != "Todas":
        filtrados = [(nome, dados) for nome, dados in alimentos.items() if dados["grupo"] in grupos]
    if not filtrados:
        return None, None

    if tipo in ("Proteína", "Proteína leve") and refeicao in ("Almoço", "Janta"):
        misturas = [(nome, dados) for nome, dados in filtrados if eh_mistura(nome, dados)]
        if misturas:
            filtrados = misturas
    elif tipo in ("Proteína", "Proteína leve"):
        proteinas_leves = [(nome, dados) for nome, dados in filtrados if not eh_mistura(nome, dados)]
        if proteinas_leves:
            filtrados = proteinas_leves

    usados = usados or {}
    return sorted(filtrados, key=lambda item: (usados.get(item[0], 0), item[1]["preco"], item[0]))[0]


def montar_plano(alimentos, objetivo, faixa):
    plano = []
    usados = {}
    for refeicao in REFEICOES:
        mult = multiplicador_refeicao(refeicao, objetivo)
        for tipo, grupos, qtd in estruturas_refeicao(refeicao, mult):
            nome, alimento = escolher_alimento(alimentos, grupos, faixa, refeicao=refeicao, tipo=tipo, usados=usados)
            if not alimento:
                continue
            usados[nome] = usados.get(nome, 0) + 1
            quantidade = max(round(qtd, 1), 0.5)
            plano.append({
                "Refeição": refeicao,
                "Tipo": tipo,
                "Alimento": nome,
                "Faixa": alimento["faixa"],
                "Porção base": alimento["porcao"],
                "Quantidade sugerida": f"{quantidade}x porção",
                "Calorias aprox.": round(alimento["calorias"] * quantidade),
                "Proteína aprox. (g)": round(alimento["proteina"] * quantidade, 1),
                "Carbo aprox. (g)": round(alimento["carbo"] * quantidade, 1),
                "Gordura aprox. (g)": round(alimento["gordura"] * quantidade, 1),
                "Custo aprox. (R$)": round(alimento["preco"] * quantidade, 2),
            })
    return plano


def montar_lista_compras(plano):
    compras = {}
    for item in plano:
        chave = (item["Alimento"], item["Porção base"])
        atual = compras.setdefault(chave, {"Custo diário (R$)": 0, "Calorias": 0})
        atual["Custo diário (R$)"] += item["Custo aprox. (R$)"]
        atual["Calorias"] += item["Calorias aprox."]

    linhas = []
    for (alimento, porcao), valores in compras.items():
        linhas.append({
            "Alimento": alimento,
            "Porção base": porcao,
            "Custo diário (R$)": round(valores["Custo diário (R$)"], 2),
            "Custo mensal (R$)": round(valores["Custo diário (R$)"] * 30, 2),
            "Calorias/dia": round(valores["Calorias"]),
        })
    return sorted(linhas, key=lambda row: row["Alimento"])


def montar_cardapio_refeicoes(plano):
    cardapio = []
    for refeicao in REFEICOES:
        itens = [item for item in plano if item["Refeição"] == refeicao]
        if not itens:
            continue

        def listar_por_tipos(tipos):
            selecionados = [item for item in itens if item["Tipo"] in tipos]
            if not selecionados:
                return "-"
            return " + ".join(
                f"{item['Alimento']} ({item['Quantidade sugerida']} - {item['Porção base']})"
                for item in selecionados
            )

        alimentos_detalhados = [
            f"{item['Alimento']} ({item['Quantidade sugerida']} - {item['Porção base']})"
            for item in itens
        ]
        cardapio.append({
            "Refeição": refeicao,
            "Prato / combinação": " + ".join(alimentos_detalhados),
            "Base / carboidrato": listar_por_tipos(["Carboidrato", "Carboidrato/Fruta", "Fruta/Carboidrato"]),
            "Acompanhamento": listar_por_tipos(["Legumes/Verduras", "Complemento", "Fruta/Gordura"]),
            "Alimentos detalhados": " + ".join(alimentos_detalhados),
            "Calorias aprox.": round(sum(item["Calorias aprox."] for item in itens)),
            "Proteína (g)": round(sum(item["Proteína aprox. (g)"] for item in itens), 1),
            "Carbo (g)": round(sum(item["Carbo aprox. (g)"] for item in itens), 1),
            "Gordura (g)": round(sum(item["Gordura aprox. (g)"] for item in itens), 1),
            "Custo aprox. (R$)": round(sum(item["Custo aprox. (R$)"] for item in itens), 2),
        })
    return cardapio


class SistemaNutricional(QMainWindow):
    def __init__(self):
        super().__init__()
        self.alimentos = carregar_alimentos()
        self.alimentos_marcados = set(self.alimentos.keys())
        self.plano_atual = []
        self.cardapio_atual = []
        self.compras_atual = []

        self.setWindowTitle("Sistema Nutricional Inteligente")
        self.resize(1380, 760)
        self.setMinimumSize(1180, 700)
        self._montar_interface()
        self.gerar_plano()

    def _montar_interface(self):
        raiz = QWidget()
        raiz.setObjectName("raiz")
        self.setCentralWidget(raiz)

        layout = QHBoxLayout(raiz)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        painel_entrada = self._painel_entrada()
        rolagem_entrada = QScrollArea()
        rolagem_entrada.setWidgetResizable(True)
        rolagem_entrada.setFrameShape(QScrollArea.NoFrame)
        rolagem_entrada.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        rolagem_entrada.setWidget(painel_entrada)
        rolagem_entrada.setMinimumWidth(410)
        rolagem_entrada.setMaximumWidth(450)
        layout.addWidget(rolagem_entrada, stretch=0)
        layout.addWidget(self._painel_resultados(), stretch=3)

        self.setStyleSheet(
            """
            QWidget#raiz { background: #f6f7fb; color: #172033; font-family: Segoe UI; font-size: 12px; }
            QWidget#sidebar { background: white; border: 1px solid #dfe3eb; border-radius: 8px; }
            QScrollArea { background: transparent; border: 0; }
            QGroupBox { background: white; border: 1px solid #dfe3eb; border-radius: 6px; margin-top: 8px; padding: 10px; font-weight: 700; }
            QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; }
            QLineEdit, QComboBox, QSpinBox, QDoubleSpinBox { min-height: 28px; padding: 2px 8px; border: 1px solid #cfd6e4; border-radius: 4px; background: white; }
            QComboBox::drop-down { width: 24px; }
            QPushButton { padding: 8px 10px; border-radius: 4px; border: 1px solid #1d7a5f; background: #1d7a5f; color: white; font-weight: 700; }
            QPushButton:hover { background: #17634d; }
            QTableWidget { background: white; border: 1px solid #dfe3eb; gridline-color: #edf0f5; }
            QHeaderView::section { background: #eef5f1; padding: 5px; border: 1px solid #dfe3eb; font-weight: 700; }
            QLabel#titulo { font-size: 23px; font-weight: 800; }
            QLabel#metrica { font-size: 18px; font-weight: 800; color: #1d7a5f; }
            """
        )

    def _painel_entrada(self):
        painel = QWidget()
        painel.setObjectName("sidebar")
        painel.setMinimumWidth(390)
        painel.setMaximumWidth(430)
        painel.setMinimumHeight(760)
        layout = QVBoxLayout(painel)
        layout.setContentsMargins(12, 10, 12, 12)
        layout.setSpacing(8)

        titulo = QLabel("Sistema Nutricional")
        titulo.setObjectName("titulo")
        layout.addWidget(titulo)

        aviso = QLabel("Ferramenta educativa. Não substitui nutricionista, médico ou orientação individualizada.")
        aviso.setWordWrap(True)
        aviso.setStyleSheet("color: #6b7280;")
        layout.addWidget(aviso)

        grupo = QGroupBox("Dados da pessoa")
        form = QFormLayout(grupo)
        form.setContentsMargins(10, 8, 10, 10)
        form.setHorizontalSpacing(10)
        form.setVerticalSpacing(6)
        form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        self.nome = QLineEdit("Fernando")
        self.idade = QSpinBox()
        self.idade.setRange(10, 100)
        self.idade.setValue(25)
        self.idade.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.peso = QDoubleSpinBox()
        self.peso.setRange(30, 300)
        self.peso.setValue(80)
        self.peso.setSuffix(" kg")
        self.peso.setDecimals(1)
        self.peso.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.altura = QSpinBox()
        self.altura.setRange(100, 250)
        self.altura.setValue(175)
        self.altura.setSuffix(" cm")
        self.altura.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.objetivo = QComboBox()
        self.objetivo.addItems(["Emagrecer", "Manter peso", "Ganhar massa"])
        self.modo = QComboBox()
        self.modo.addItems(["Automático por faixa de preço", "Personalizado com alimentos em casa"])
        self.modo.currentTextChanged.connect(self.atualizar_modo)
        self.faixa = QComboBox()
        self.faixa.addItems(["Econômica", "Média", "Premium", "Todas"])
        campos = [self.nome, self.idade, self.peso, self.altura, self.objetivo, self.modo, self.faixa]
        for campo in campos:
            campo.setMinimumWidth(0)
            campo.setMinimumHeight(30)
        for combo in [self.objetivo, self.modo, self.faixa]:
            combo.setMinimumContentsLength(18)

        for rotulo, campo in [
            ("Nome", self.nome),
            ("Idade", self.idade),
            ("Peso", self.peso),
            ("Altura", self.altura),
            ("Objetivo", self.objetivo),
            ("Modo", self.modo),
            ("Faixa de preço", self.faixa),
        ]:
            form.addRow(rotulo, campo)
        layout.addWidget(grupo)

        self.alimentos_box = QGroupBox("Alimentos disponíveis")
        alimentos_layout = QVBoxLayout(self.alimentos_box)
        alimentos_layout.setContentsMargins(10, 8, 10, 10)
        alimentos_layout.setSpacing(7)
        self.texto_alimentos = QLabel("No modo personalizado, marque apenas os alimentos que você tem em casa.")
        self.texto_alimentos.setWordWrap(True)
        self.texto_alimentos.setStyleSheet("color: #6b7280; font-weight: 400;")
        alimentos_layout.addWidget(self.texto_alimentos)

        self.lista_alimentos = QListWidget()
        self.lista_alimentos.itemChanged.connect(self.atualizar_marcacao_alimento)
        self.lista_alimentos.currentItemChanged.connect(self.carregar_alimento_selecionado)
        self.busca_alimento = QLineEdit()
        self.busca_alimento.setPlaceholderText("Buscar alimento pelo nome")
        self.busca_alimento.textChanged.connect(self.atualizar_lista_alimentos)
        alimentos_layout.addWidget(self.busca_alimento)
        self.filtro_grupo_alimento = QComboBox()
        self.filtro_grupo_alimento.addItems(["Todos", "Carboidrato", "Fruta", "Proteína", "Gordura", "Legume/Verdura"])
        self.filtro_grupo_alimento.currentTextChanged.connect(self.atualizar_lista_alimentos)
        alimentos_layout.addWidget(self.filtro_grupo_alimento)
        self.atualizar_lista_alimentos()
        self.lista_alimentos.setMinimumHeight(140)
        alimentos_layout.addWidget(self.lista_alimentos)

        preco_layout = QHBoxLayout()
        preco_layout.setSpacing(8)
        self.preco_alimento = QDoubleSpinBox()
        self.preco_alimento.setRange(0, 999)
        self.preco_alimento.setDecimals(2)
        self.preco_alimento.setPrefix("R$ ")
        self.preco_alimento.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.preco_alimento.setMinimumWidth(120)
        btn_preco = QPushButton("Atualizar preço")
        btn_preco.clicked.connect(self.atualizar_preco_alimento)
        preco_layout.addWidget(self.preco_alimento)
        preco_layout.addWidget(btn_preco)
        alimentos_layout.addLayout(preco_layout)

        botoes_alimentos = QHBoxLayout()
        botoes_alimentos.setSpacing(8)
        btn_todos = QPushButton("Marcar todos")
        btn_todos.clicked.connect(lambda: self.definir_alimentos(Qt.Checked))
        btn_limpar = QPushButton("Desmarcar")
        btn_limpar.clicked.connect(lambda: self.definir_alimentos(Qt.Unchecked))
        botoes_alimentos.addWidget(btn_todos)
        botoes_alimentos.addWidget(btn_limpar)
        alimentos_layout.addLayout(botoes_alimentos)

        btn_adicionar = QPushButton("Adicionar novo alimento")
        btn_adicionar.clicked.connect(self.abrir_dialogo_adicionar_alimento)
        alimentos_layout.addWidget(btn_adicionar)
        layout.addWidget(self.alimentos_box)

        btn_gerar = QPushButton("Gerar plano")
        btn_gerar.clicked.connect(self.gerar_plano)
        layout.addWidget(btn_gerar)

        botoes_exportacao = QHBoxLayout()
        btn_exportar = QPushButton("Exportar Excel")
        btn_exportar.clicked.connect(self.exportar_excel)
        btn_pdf = QPushButton("Exportar PDF")
        btn_pdf.clicked.connect(self.exportar_pdf)
        botoes_exportacao.addWidget(btn_exportar)
        botoes_exportacao.addWidget(btn_pdf)
        layout.addLayout(botoes_exportacao)
        self.atualizar_modo()
        return painel

    def _painel_resultados(self):
        painel = QWidget()
        layout = QVBoxLayout(painel)
        layout.setSpacing(10)

        self.metricas = {}
        metricas_grid = QGridLayout()
        metricas_grid.setHorizontalSpacing(10)
        metricas_grid.setVerticalSpacing(6)
        for indice, chave in enumerate(["IMC", "Classificação", "Calorias", "Água", "Proteína", "Carboidrato", "Gordura", "Custo mensal"]):
            card = QGroupBox(chave)
            card_layout = QVBoxLayout(card)
            card_layout.setContentsMargins(10, 8, 10, 8)
            valor = QLabel("-")
            valor.setObjectName("metrica")
            card_layout.addWidget(valor)
            self.metricas[chave] = valor
            metricas_grid.addWidget(card, indice // 4, indice % 4)
        layout.addLayout(metricas_grid)

        filtros_box = QGroupBox("Filtros do cardápio")
        filtros_layout = QHBoxLayout(filtros_box)
        self.filtro_refeicao = QComboBox()
        self.filtro_refeicao.addItems(["Todas"] + REFEICOES)
        self.filtro_prato = QLineEdit()
        self.filtro_prato.setPlaceholderText("Filtrar prato/alimento")
        self.filtro_mistura = QLineEdit()
        self.filtro_mistura.setPlaceholderText("Filtrar carne/frango/ovo")
        btn_limpar_filtros = QPushButton("Limpar filtros")
        btn_limpar_filtros.clicked.connect(self.limpar_filtros_cardapio)
        self.filtro_refeicao.currentTextChanged.connect(self.aplicar_filtros_cardapio)
        self.filtro_prato.textChanged.connect(self.aplicar_filtros_cardapio)
        self.filtro_mistura.textChanged.connect(self.aplicar_filtros_cardapio)
        filtros_layout.addWidget(self.filtro_refeicao)
        filtros_layout.addWidget(self.filtro_prato)
        filtros_layout.addWidget(self.filtro_mistura)
        filtros_layout.addWidget(btn_limpar_filtros)
        layout.addWidget(filtros_box)

        self.abas = QTabWidget()
        self.tabela_cardapio = self._nova_tabela()
        self.tabela_plano = self._nova_tabela()
        self.tabela_alimentos = self._nova_tabela()
        self.tabela_compras = self._nova_tabela()
        self.abas.addTab(self.tabela_cardapio, "Cardápio por refeição")
        self.abas.addTab(self.tabela_plano, "Detalhes nutricionais")
        self.abas.addTab(self.tabela_compras, "Lista de compras")
        self.abas.addTab(self.tabela_alimentos, "Base de alimentos")
        layout.addWidget(self.abas, stretch=1)
        return painel

    def _nova_tabela(self):
        tabela = QTableWidget()
        tabela.setAlternatingRowColors(True)
        tabela.setEditTriggers(QTableWidget.NoEditTriggers)
        tabela.setSelectionBehavior(QTableWidget.SelectRows)
        return tabela

    def atualizar_modo(self):
        personalizado = self.modo.currentText() == "Personalizado com alimentos em casa"
        self.faixa.setEnabled(not personalizado)
        if personalizado:
            self.alimentos_box.setTitle("Alimentos que tenho em casa")
            self.texto_alimentos.setText("Marque apenas os alimentos disponíveis. O plano será montado somente com esses itens.")
        else:
            self.alimentos_box.setTitle("Base de alimentos")
            self.texto_alimentos.setText("No modo automático, o sistema usa a base completa e respeita a faixa de preço selecionada.")

    def definir_alimentos(self, estado):
        for indice in range(self.lista_alimentos.count()):
            self.lista_alimentos.item(indice).setCheckState(estado)

    def atualizar_lista_alimentos(self):
        termo = self.busca_alimento.text().strip().lower() if hasattr(self, "busca_alimento") else ""
        grupo = self.filtro_grupo_alimento.currentText() if hasattr(self, "filtro_grupo_alimento") else "Todos"
        self.lista_alimentos.blockSignals(True)
        self.lista_alimentos.clear()
        for nome, dados in sorted(self.alimentos.items()):
            if termo and termo not in nome.lower():
                continue
            if grupo != "Todos" and dados["grupo"] != grupo:
                continue
            item = QListWidgetItem(f"{nome} - {dados['grupo']} - {dados['faixa']} - R$ {dados['preco']:.2f}")
            item.setData(Qt.UserRole, nome)
            item.setCheckState(Qt.Checked if nome in self.alimentos_marcados else Qt.Unchecked)
            self.lista_alimentos.addItem(item)
        self.lista_alimentos.blockSignals(False)

    def atualizar_marcacao_alimento(self, item):
        nome = item.data(Qt.UserRole)
        if item.checkState() == Qt.Checked:
            self.alimentos_marcados.add(nome)
        else:
            self.alimentos_marcados.discard(nome)

    def carregar_alimento_selecionado(self, item, _anterior=None):
        if not item or not hasattr(self, "preco_alimento"):
            return
        nome = item.data(Qt.UserRole)
        self.preco_alimento.setValue(float(self.alimentos[nome]["preco"]))

    def atualizar_preco_alimento(self):
        item = self.lista_alimentos.currentItem()
        if not item:
            QMessageBox.information(self, "Selecione um alimento", "Selecione um alimento para atualizar o preço.")
            return
        nome = item.data(Qt.UserRole)
        self.alimentos[nome]["preco"] = round(self.preco_alimento.value(), 2)
        salvar_alimentos_personalizados(self.alimentos)
        self.atualizar_lista_alimentos()
        self.gerar_plano()

    def abrir_dialogo_adicionar_alimento(self):
        dialog = QDialog(self)
        dialog.setWindowTitle("Adicionar alimento")
        dialog.setMinimumWidth(520)
        layout = QVBoxLayout(dialog)
        form = QFormLayout()
        form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        form.setVerticalSpacing(10)

        nome = QLineEdit()
        grupo = QComboBox()
        grupo.addItems(["Carboidrato", "Fruta", "Proteína", "Gordura", "Legume/Verdura"])
        faixa = QComboBox()
        faixa.addItems(["Econômica", "Média", "Premium"])
        porcao = QLineEdit("100g")
        calorias = QDoubleSpinBox()
        calorias.setRange(0, 2000)
        proteina = QDoubleSpinBox()
        proteina.setRange(0, 300)
        carbo = QDoubleSpinBox()
        carbo.setRange(0, 300)
        gordura = QDoubleSpinBox()
        gordura.setRange(0, 300)
        preco = QDoubleSpinBox()
        preco.setRange(0, 999)
        preco.setDecimals(2)
        preco.setPrefix("R$ ")

        for campo in [calorias, proteina, carbo, gordura, preco]:
            campo.setButtonSymbols(QAbstractSpinBox.NoButtons)
            campo.setMinimumHeight(34)

        form.addRow("Nome", nome)
        form.addRow("Grupo", grupo)
        form.addRow("Faixa", faixa)
        form.addRow("Porção", porcao)
        form.addRow("Calorias", calorias)
        form.addRow("Proteína", proteina)
        form.addRow("Carbo", carbo)
        form.addRow("Gordura", gordura)
        form.addRow("Preço", preco)
        layout.addLayout(form)

        botoes = QHBoxLayout()
        btn_cancelar = QPushButton("Cancelar")
        btn_cancelar.clicked.connect(dialog.reject)
        btn_salvar = QPushButton("Adicionar na base")
        btn_salvar.clicked.connect(dialog.accept)
        botoes.addWidget(btn_cancelar)
        botoes.addWidget(btn_salvar)
        layout.addLayout(botoes)

        if dialog.exec() != QDialog.Accepted:
            return

        self.adicionar_alimento(
            nome.text(),
            grupo.currentText(),
            faixa.currentText(),
            porcao.text(),
            calorias.value(),
            proteina.value(),
            carbo.value(),
            gordura.value(),
            preco.value(),
        )

    def adicionar_alimento(self, nome, grupo, faixa, porcao, calorias, proteina, carbo, gordura, preco):
        nome = nome.strip()
        if not nome:
            QMessageBox.warning(self, "Nome obrigatório", "Informe o nome do alimento.")
            return
        self.alimentos[nome] = {
            "grupo": grupo,
            "faixa": faixa,
            "porcao": porcao.strip() or "100g",
            "calorias": round(calorias, 1),
            "proteina": round(proteina, 1),
            "carbo": round(carbo, 1),
            "gordura": round(gordura, 1),
            "preco": round(preco, 2),
        }
        self.alimentos_marcados.add(nome)
        salvar_alimentos_personalizados(self.alimentos)
        self.atualizar_lista_alimentos()
        QMessageBox.information(self, "Alimento adicionado", f"{nome} foi adicionado à base.")

    def alimentos_selecionados(self):
        escolhidos = {}
        for indice in range(self.lista_alimentos.count()):
            item = self.lista_alimentos.item(indice)
            if item.checkState() == Qt.Checked:
                nome = item.data(Qt.UserRole)
                escolhidos[nome] = self.alimentos[nome]
        return escolhidos

    def gerar_plano(self):
        personalizado = self.modo.currentText() == "Personalizado com alimentos em casa"
        alimentos = self.alimentos_selecionados() if personalizado else self.alimentos
        faixa = "Todas" if personalizado else self.faixa.currentText()
        if not alimentos:
            QMessageBox.warning(self, "Sem alimentos", "Selecione ao menos alguns alimentos para montar o plano.")
            return

        peso = self.peso.value()
        objetivo = self.objetivo.currentText()
        imc = calcular_imc(peso, self.altura.value())
        proteina, carbo, gordura = calcular_macros(peso, objetivo)
        calorias = calcular_calorias(peso, objetivo)
        agua = peso * 35 / 1000

        self.plano_atual = montar_plano(alimentos, objetivo, faixa)
        if not self.plano_atual:
            QMessageBox.warning(
                self,
                "Plano incompleto",
                "Não foi possível montar o plano. Selecione alimentos de grupos diferentes, como proteína, carboidrato, fruta e legumes.",
            )
            return
        self.compras_atual = montar_lista_compras(self.plano_atual)
        self.cardapio_atual = montar_cardapio_refeicoes(self.plano_atual)
        custo_diario = sum(item["Custo aprox. (R$)"] for item in self.plano_atual)

        self.metricas["IMC"].setText(f"{imc:.1f}")
        self.metricas["Classificação"].setText(classificar_imc(imc))
        self.metricas["Calorias"].setText(f"{calorias:.0f} kcal/dia")
        self.metricas["Água"].setText(f"{agua:.1f} L/dia")
        self.metricas["Proteína"].setText(f"{proteina} g")
        self.metricas["Carboidrato"].setText(f"{carbo} g")
        self.metricas["Gordura"].setText(f"{gordura} g")
        self.metricas["Custo mensal"].setText(f"R$ {custo_diario * 30:.2f}")

        self.aplicar_filtros_cardapio()
        self._preencher_tabela(self.tabela_plano, self.plano_atual)
        self._preencher_tabela(self.tabela_compras, self.compras_atual)
        self._preencher_tabela(self.tabela_alimentos, self.base_alimentos_linhas())

    def base_alimentos_linhas(self):
        return [
            {
                "Alimento": nome,
                "Grupo": dados["grupo"],
                "Faixa": dados["faixa"],
                "Porção": dados["porcao"],
                "Calorias": dados["calorias"],
                "Proteína (g)": dados["proteina"],
                "Carbo (g)": dados["carbo"],
                "Gordura (g)": dados["gordura"],
                "Preço (R$)": dados["preco"],
            }
            for nome, dados in sorted(self.alimentos.items())
        ]

    def limpar_filtros_cardapio(self):
        self.filtro_refeicao.setCurrentText("Todas")
        self.filtro_prato.clear()
        self.filtro_mistura.clear()
        self.aplicar_filtros_cardapio()

    def aplicar_filtros_cardapio(self):
        if not hasattr(self, "tabela_cardapio"):
            return
        refeicao = self.filtro_refeicao.currentText()
        prato = self.filtro_prato.text().strip().lower()
        mistura = self.filtro_mistura.text().strip().lower()
        linhas = []
        for linha in self.cardapio_atual:
            if refeicao != "Todas" and linha["Refeição"] != refeicao:
                continue
            if prato and prato not in linha["Prato / combinação"].lower() and prato not in linha["Alimentos detalhados"].lower():
                continue
            if mistura and mistura not in linha["Prato / combinação"].lower():
                continue
            linhas.append(linha)
        self._preencher_tabela(self.tabela_cardapio, linhas)

    def _preencher_tabela(self, tabela, linhas):
        tabela.clear()
        if not linhas:
            tabela.setRowCount(0)
            tabela.setColumnCount(0)
            return

        colunas = list(linhas[0].keys())
        tabela.setColumnCount(len(colunas))
        tabela.setRowCount(len(linhas))
        tabela.setHorizontalHeaderLabels(colunas)

        for linha_idx, linha in enumerate(linhas):
            cor = QColor("#ffffff")
            if "Refeição" in linha:
                if linha["Refeição"] in ("Almoço", "Janta"):
                    cor = QColor("#eef8f1")
                elif "Lanche" in linha["Refeição"]:
                    cor = QColor("#fff8e6")
            for coluna_idx, coluna in enumerate(colunas):
                item = QTableWidgetItem(str(linha[coluna]))
                item.setBackground(cor)
                if isinstance(linha[coluna], (int, float)):
                    item.setTextAlignment(Qt.AlignCenter)
                if coluna in ("Prato / combinação", "Base / carboidrato", "Acompanhamento", "Alimentos detalhados"):
                    item.setToolTip(str(linha[coluna]))
                tabela.setItem(linha_idx, coluna_idx, item)

        tabela.resizeColumnsToContents()
        tabela.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        if "Prato / combinação" in colunas:
            tabela.horizontalHeader().setSectionResizeMode(colunas.index("Prato / combinação"), QHeaderView.Stretch)
        elif tabela.columnCount() > 2:
            tabela.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)

    def exportar_excel(self):
        if not self.plano_atual:
            QMessageBox.information(self, "Sem plano", "Gere um plano antes de exportar.")
            return

        nome_sugerido = f"plano_nutricional_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        caminho, _ = QFileDialog.getSaveFileName(self, "Salvar plano", str(APP_DIR / nome_sugerido), "Excel (*.xlsx)")
        if not caminho:
            return

        wb = Workbook()
        self._adicionar_aba_excel(wb.active, "Cardápio por refeição", self.cardapio_atual)
        plano = wb.create_sheet("Detalhes nutricionais")
        self._adicionar_aba_excel(plano, "Detalhes nutricionais", self.plano_atual)
        compras = wb.create_sheet("Lista de compras")
        self._adicionar_aba_excel(compras, "Lista de compras", self.compras_atual)
        alimentos = wb.create_sheet("Base de alimentos")
        self._adicionar_aba_excel(alimentos, "Base de alimentos", self.base_alimentos_linhas())
        wb.save(caminho)
        QMessageBox.information(self, "Exportação concluída", "Plano exportado com sucesso.")

    def exportar_pdf(self):
        if not self.plano_atual:
            QMessageBox.information(self, "Sem plano", "Gere um plano antes de exportar.")
            return

        nome_sugerido = f"plano_nutricional_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        caminho, _ = QFileDialog.getSaveFileName(self, "Salvar relatório PDF", str(APP_DIR / nome_sugerido), "PDF (*.pdf)")
        if not caminho:
            return

        writer = QPdfWriter(caminho)
        writer.setPageSize(QPageSize(QPageSize.A4))
        writer.setPageMargins(QMarginsF(16, 16, 16, 16))

        documento = QTextDocument()
        documento.setPageSize(writer.pageLayout().paintRectPoints().size())
        documento.setHtml(self._html_relatorio_pdf_legivel())
        documento.print_(writer)
        QMessageBox.information(self, "Exportação concluída", "Relatório PDF exportado com sucesso.")

    def _html_relatorio(self):
        nome = escape(self.nome.text().strip() or "Paciente")
        modo = escape(self.modo.currentText())
        objetivo = escape(self.objetivo.currentText())
        data = datetime.now().strftime("%d/%m/%Y %H:%M")
        metricas = "".join(
            f"<div class='metric'><span>{escape(titulo)}</span><strong>{escape(label.text())}</strong></div>"
            for titulo, label in self.metricas.items()
        )

        return f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; color: #172033; }}
                h1 {{ color: #1d7a5f; margin-bottom: 4px; }}
                h2 {{ color: #1d7a5f; margin-top: 22px; }}
                .muted {{ color: #667085; font-size: 11px; }}
                .metrics {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 8px; margin: 14px 0; }}
                .metric {{ border: 1px solid #dfe3eb; border-radius: 6px; padding: 8px; }}
                .metric span {{ display: block; color: #667085; font-size: 10px; }}
                .metric strong {{ font-size: 14px; }}
                table {{ width: 100%; border-collapse: collapse; font-size: 9px; }}
                th {{ background: #eef5f1; }}
                th, td {{ border: 1px solid #dfe3eb; padding: 4px; text-align: left; }}
                .warning {{ margin-top: 18px; color: #667085; font-size: 10px; }}
            </style>
        </head>
        <body>
            <h1>Plano Nutricional</h1>
            <div class="muted">Pessoa: {nome} | Objetivo: {objetivo} | Modo: {modo} | Gerado em {data}</div>
            <div class="metrics">{metricas}</div>
            <h2>Cardápio por refeição</h2>
            {self._html_tabela(self.cardapio_atual)}
            <h2>Detalhes nutricionais</h2>
            {self._html_tabela(self.plano_atual)}
            <h2>Lista de compras estimada</h2>
            {self._html_tabela(self.compras_atual)}
            <p class="warning">Ferramenta educativa. Não substitui nutricionista, médico ou orientação individualizada.</p>
        </body>
        </html>
        """

    def _html_tabela(self, linhas):
        if not linhas:
            return "<p>Sem dados.</p>"
        colunas = list(linhas[0].keys())
        cabecalho = "".join(f"<th>{escape(coluna)}</th>" for coluna in colunas)
        corpo = []
        for linha in linhas:
            celulas = "".join(f"<td>{escape(str(linha[coluna]))}</td>" for coluna in colunas)
            corpo.append(f"<tr>{celulas}</tr>")
        return f"<table><thead><tr>{cabecalho}</tr></thead><tbody>{''.join(corpo)}</tbody></table>"

    def _html_relatorio_pdf_legivel(self):
        nome = escape(self.nome.text().strip() or "Paciente")
        modo = escape(self.modo.currentText())
        objetivo = escape(self.objetivo.currentText())
        data = datetime.now().strftime("%d/%m/%Y %H:%M")
        return f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; color: #172033; font-size: 11pt; }}
                h1 {{ color: #1d7a5f; margin-bottom: 4px; }}
                h2 {{ color: #1d7a5f; margin-top: 18px; margin-bottom: 6px; }}
                .muted {{ color: #667085; font-size: 9pt; }}
                table {{ width: 100%; border-collapse: collapse; font-size: 9.5pt; margin-bottom: 10px; }}
                th {{ background: #eef5f1; }}
                th, td {{ border: 1px solid #dfe3eb; padding: 5px; text-align: left; vertical-align: top; }}
                .meal {{ border: 1px solid #dfe3eb; padding: 8px; margin-bottom: 8px; }}
                .meal-title {{ font-weight: bold; color: #1d7a5f; font-size: 12pt; }}
                .warning {{ margin-top: 18px; color: #667085; font-size: 9pt; }}
                .page-break {{ page-break-before: always; }}
            </style>
        </head>
        <body>
            <h1>Plano Nutricional</h1>
            <div class="muted">Pessoa: {nome} | Objetivo: {objetivo} | Modo: {modo} | Gerado em {data}</div>
            {self._html_metricas_pdf()}
            <h2>Cardápio por refeição</h2>
            {self._html_cardapio_pdf()}
            <h2>Detalhes nutricionais</h2>
            {self._html_detalhes_pdf()}
            <h2>Lista de compras estimada</h2>
            {self._html_compras_pdf()}
            <div class="page-break"></div>
            <h2>Base de alimentos usada</h2>
            {self._html_base_alimentos_pdf()}
            <p class="warning">Ferramenta educativa. Não substitui nutricionista, médico ou orientação individualizada.</p>
        </body>
        </html>
        """

    def _html_metricas_pdf(self):
        linhas = ""
        pares = list(self.metricas.items())
        for indice in range(0, len(pares), 4):
            celulas = "".join(
                f"<td><strong>{escape(titulo)}</strong><br>{escape(label.text())}</td>"
                for titulo, label in pares[indice:indice + 4]
            )
            linhas += f"<tr>{celulas}</tr>"
        return f"<table>{linhas}</table>"

    def _html_cardapio_pdf(self):
        blocos = []
        for linha in self.cardapio_atual:
            blocos.append(
                "<div class='meal'>"
                f"<div class='meal-title'>{escape(linha['Refeição'])}</div>"
                f"<p><strong>Prato:</strong> {escape(linha['Prato / combinação'])}</p>"
                f"<p><strong>Base:</strong> {escape(linha['Base / carboidrato'])}</p>"
                f"<p><strong>Acompanhamento:</strong> {escape(linha['Acompanhamento'])}</p>"
                f"<p><strong>Resumo:</strong> {linha['Calorias aprox.']} kcal | "
                f"Proteína {linha['Proteína (g)']}g | Carbo {linha['Carbo (g)']}g | "
                f"Gordura {linha['Gordura (g)']}g | R$ {linha['Custo aprox. (R$)']:.2f}</p>"
                "</div>"
            )
        return "".join(blocos) if blocos else "<p>Sem dados.</p>"

    def _html_detalhes_pdf(self):
        colunas = ["Refeição", "Alimento", "Quantidade sugerida", "Calorias aprox.", "Proteína aprox. (g)", "Carbo aprox. (g)", "Gordura aprox. (g)"]
        return self._html_tabela_colunas(colunas, self.plano_atual)

    def _html_compras_pdf(self):
        colunas = ["Alimento", "Porção base", "Custo diário (R$)", "Custo mensal (R$)"]
        return self._html_tabela_colunas(colunas, self.compras_atual)

    def _html_base_alimentos_pdf(self):
        colunas = ["Alimento", "Grupo", "Faixa", "Porção", "Preço (R$)"]
        return self._html_tabela_colunas(colunas, self.base_alimentos_linhas())

    def _html_tabela_colunas(self, colunas, linhas):
        if not linhas:
            return "<p>Sem dados.</p>"
        cabecalho = "".join(f"<th>{escape(coluna)}</th>" for coluna in colunas)
        corpo = []
        for linha in linhas:
            celulas = "".join(f"<td>{escape(str(linha.get(coluna, '')))}</td>" for coluna in colunas)
            corpo.append(f"<tr>{celulas}</tr>")
        return f"<table><thead><tr>{cabecalho}</tr></thead><tbody>{''.join(corpo)}</tbody></table>"

    def _adicionar_aba_excel(self, sheet, titulo, linhas):
        sheet.title = titulo[:31]
        if not linhas:
            return
        colunas = list(linhas[0].keys())
        sheet.append(colunas)
        for linha in linhas:
            sheet.append([linha[coluna] for coluna in colunas])
        for coluna in sheet.columns:
            largura = max(len(str(celula.value or "")) for celula in coluna) + 2
            sheet.column_dimensions[coluna[0].column_letter].width = min(largura, 42)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    janela = SistemaNutricional()
    janela.show()
    sys.exit(app.exec())
